import os

from flask import Flask
from flask import redirect, render_template, url_for

# import openpyxl module
import openpyxl


def create_app(test_config=None):
    # Give the location of the file
    path = "TKH Student Projects.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    cell_obj = sheet_obj.cell(row = 1, column = 1)
    max_col_num = sheet_obj.max_column
    max_row_num = sheet_obj.max_row

    print(max_col_num)
    # col_content = {1: "Student Name", 2:"Track", 3:"Project Title", 4:"Link to Project", 5:"Description of Project", 6: "Github Link", 7: "Photo Link"}
    col_content = {}
    for c in range(1, max_col_num+1):
        col_content[c] = sheet_obj.cell(row = 1, column = c).value
    


    # create and configure the app
    app = Flask(__name__, instance_relative_config=True)
    app.config.from_mapping(
        SECRET_KEY='dev',
        DATABASE=os.path.join(app.instance_path, 'flaskr.sqlite'),
    )

    if test_config is None:
        # load the instance config, if it exists, when not testing
        app.config.from_pyfile('config.py', silent=True)
    else:
        # load the test config if passed in
        app.config.from_mapping(test_config)

    # ensure the instance folder exists
    try:
        os.makedirs(app.instance_path)
    except OSError:
        pass

    track_info = {}
    track_keyword_website = {}
    # read project urls from "TKH Student Projects.xlsx"
    for r in range(2, max_row_num):
        track = sheet_obj.cell(row = r, column = 2).value
        student_first_name = sheet_obj.cell(row = r, column = 1).value.split(" ")[0]
        keyword = sheet_obj.cell(row = r, column = 3).value.split(" ")[0] if " " in sheet_obj.cell(row = r, column = 3).value else sheet_obj.cell(row = r, column = 3).value
        keyword_name = keyword + "_" + student_first_name
        website = sheet_obj.cell(row = r, column = 4).value
        if track not in track_keyword_website:
            track_keyword_website[track] = {}
        track_keyword_website[track].update({keyword_name:website})

        if track not in track_info:
            track_info[track] = []
        project_info = {}
        for c in range(1, max_col_num+1):
            column_name = col_content[c]
            project_info.update({column_name: sheet_obj.cell(row = r, column = c).value}) 
        track_info[track].append(project_info)
        
        
    print (track_info)

    #track_keyword_studentname
    #http://127.0.0.1:5000/projects/WD/Bloom_Quiana
    #http://127.0.0.1:5000/projects/D/Treasury_Gyasi
    #http://127.0.0.1:5000/projects/D/Food_Jessica

    @app.route('/projects/<track>/<keyword_name>')
    def endpoint_detail(track, keyword_name):
        return redirect(track_keyword_website[track][keyword_name])

    #http://127.0.0.1:5000/projects/WD
    #http://127.0.0.1:5000/projects/D
    @app.route('/projects/<track>')
    def endpoint_track(track):
        return {track:track_info[track]}

    return app